import os
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    import xlsxwriter  # noqa: F401  提速 Excel 写入
    _HAS_XLSXWRITER = True
except ImportError:
    _HAS_XLSXWRITER = False

from models.analysis import (
    analyze_bottleneck,
    analyze_bottleneck_machines,
    analyze_status,
    analyze_status_derived,
    analyze_steps,
    analyze_steps_sa,
    build_cycles_df,
    build_gantt_rows,
    build_gantt_rows_sa,
    detect_units_per_tray,
    detect_tray_stats,
    down_pareto,
    measure_tray_change,
    parse_em_production,
    summarize_alarms,
    summarize_eff_coretech,
    summarize_uph,
    summarize_uph_ame,
)
from models.reason_codes import load_reason_codes
from utils.file_utils import read_files, clean_for_excel


class LogModel:
    # 文档合并的行数上限：超过后放弃合并，改为按日志文件逐个导出 Excel
    MERGE_MAX_TOTAL_ROWS = 1_000_000

    @staticmethod
    def _load_reason_map(reason_device):
        """按制程名加载 EReason 清单映射；未指定时返回 None。"""
        if not reason_device:
            return None
        return load_reason_codes(reason_device)

    # ---------- 通用：读取 / 导出 ----------
    def _read_all(self, source_dir, progress_callback=None, file_filters=None, cancel_event=None):
        def on_read_progress(frac):
            if progress_callback:
                progress_callback(10 + int(19 * frac))

        all_data = read_files(source_dir, progress_callback=on_read_progress, file_filters=file_filters,
                              cancel_event=cancel_event)
        if not all_data:
            raise ValueError("日志文件内容均为空，无法解析")
        return all_data

    def load_rows(self, source_dir, file_filters=None, progress_callback=None, cancel_event=None):
        """读取日志行（供一键分析共享，避免每项重复读取大文件）。"""
        return self._read_all(source_dir, progress_callback, file_filters=file_filters,
                              cancel_event=cancel_event)

    @staticmethod
    def _write_sheets(out_path, sheets, cancel_event=None, sheet_highlights=None):
        """写多 sheet Excel；sheet_highlights: {表名: [DataFrame 0-based 行号]}。

        xlsxwriter 引擎下直接在写表阶段用 set_row 标红（避免 openpyxl 二次加载大表），
        返回是否已应用标红；未应用时（如 openpyxl 引擎）由调用方走 _apply_highlights 兜底。
        """
        max_rows = 1048575  # Excel 单表最大行数（含表头行）
        heavy = False  # 是否包含超大表（>20万行）：跳过整体格式化，避免重新加载保存拖慢导出
        writer_kwargs = {}
        if _HAS_XLSXWRITER:
            writer_kwargs = {
                'engine': 'xlsxwriter',
                'engine_kwargs': {'options': {'nan_inf_to_errors': True}},
            }
        applied = False
        with pd.ExcelWriter(out_path, **writer_kwargs) as writer:
            for sheet_name, df in sheets.items():
                if cancel_event is not None and cancel_event.is_set():
                    raise OperationCancelled()
                hl = (sheet_highlights or {}).get(sheet_name) or []
                if len(df) <= max_rows:
                    if len(df) > LogModel.MAX_FORMAT_ROWS:
                        heavy = True
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    if hl and _HAS_XLSXWRITER:
                        LogModel._apply_row_fills(writer, sheet_name, hl, offset=0, n_rows=len(df))
                        applied = True
                else:
                    # 超限自动拆分：AllLogs → AllLogs_1, AllLogs_2 ...
                    heavy = True
                    n = (len(df) + max_rows - 1) // max_rows
                    for i in range(n):
                        if cancel_event is not None and cancel_event.is_set():
                            raise OperationCancelled()
                        chunk = df.iloc[i * max_rows:(i + 1) * max_rows]
                        name = f"{sheet_name}_{i + 1}"[:31]
                        chunk.to_excel(writer, sheet_name=name, index=False)
                        if hl and _HAS_XLSXWRITER:
                            chunk_hl = [r for r in hl if i * max_rows <= r < (i + 1) * max_rows]
                            if chunk_hl:
                                LogModel._apply_row_fills(writer, name, chunk_hl,
                                                          offset=i * max_rows, n_rows=len(chunk))
                                applied = True
        if not heavy:
            LogModel._format_workbook(out_path)
        return applied

    @staticmethod
    def _apply_row_fills(writer, sheet_name, df_rows, offset=0, n_rows=None):
        """xlsxwriter 阶段按整行标红：df_rows 为 DataFrame 0-based 行号，offset 为当前块起点。"""
        ws = writer.sheets[sheet_name]
        red = writer.book.add_format({'bg_color': 'FFC7CE'})
        if n_rows is None:
            n_rows = len(df_rows) + offset
        for r in df_rows:
            excel_row = (r - offset) + 2  # 表头占第 1 行
            if 1 <= excel_row <= n_rows + 1:
                ws.set_row(excel_row - 1, None, red)  # xlsxwriter 行号从 0 开始

    # 超大表逐格样式上限（超过则只做表头/列宽，避免卡死）
    MAX_FORMAT_ROWS = 200000

    @staticmethod
    def _format_workbook(path):
        """统一 Excel 样式：列宽自适应、内容居中、有内容加框线、表头浅蓝加粗。"""
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal='center', vertical='center')
        header_fill = PatternFill('solid', fgColor='DDEBF7')  # 浅蓝色
        header_font = Font(bold=True)

        wb = load_workbook(path)
        for ws in wb.worksheets:
            if ws.title == "步骤甘特图":
                LogModel._format_gantt_sheet(ws, center, border)
                continue
            # 表头：浅蓝填充 + 加粗 + 居中
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center

            # 列宽自适应（中文按 2 字符宽计）
            for col_cells in ws.columns:
                letter = get_column_letter(col_cells[0].column)
                max_len = 0
                for cell in col_cells:
                    if cell.value is None:
                        continue
                    length = sum(2 if ord(ch) > 127 else 1 for ch in str(cell.value))
                    if length > max_len:
                        max_len = length
                ws.column_dimensions[letter].width = min(max(max_len + 2, 8), 60)

            # 超大表跳过逐格样式，避免导出卡死
            if ws.max_row > LogModel.MAX_FORMAT_ROWS:
                continue

            # 内容单元格：居中 + 有内容加框线（空单元格保持无格线）
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.value is None or str(cell.value) == '':
                        continue
                    cell.alignment = center
                    cell.border = border
        wb.save(path)

    @staticmethod
    def _format_gantt_sheet(ws, center, border):
        """
        步骤甘特图：每行一个步骤，每列一个时间块（默认 0.1s，按数据自动放大分辨率），
        用颜色填充的列数表示该步骤时长；并行步骤在同一时间区间的多行并排可见。
        层级颜色：循环=浅蓝、批次=浅橙、盘=浅绿。
        """
        import math
        header_fill = PatternFill('solid', fgColor='DDEBF7')
        header_font = Font(bold=True)
        layer_fill = {
            '循环': PatternFill('solid', fgColor='DDEBF7'),   # 浅蓝
            '批次': PatternFill('solid', fgColor='FCE4D6'),   # 浅橙
            '盘': PatternFill('solid', fgColor='E2EFDA'),     # 浅绿
        }
        # 读取已写入的步骤数据（单元/步骤/开始秒/时长秒/结束秒/层级）
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return
        data = [r for r in rows[1:] if r[0] is not None]
        # 按表头定位列，避免列顺序错位（开始秒/结束秒/时长秒）
        header = list(rows[0])
        try:
            col_start = header.index('开始秒')
            col_end = header.index('结束秒')
        except ValueError:
            return
        col_layer = header.index('层级') if '层级' in header else 5
        max_end = max((float(r[col_end]) for r in data
                       if r[col_end] not in (None, '')), default=0.0)
        # 时间块分辨率：默认 0.1s，目标 60~120 列，超出则放大到 0.2/0.5/1/2/5/10...
        col_sec = 0.1
        if max_end > 0 and max_end / col_sec > 120:
            base = max_end / 100.0
            for step in (0.1, 0.2, 0.5, 1, 2, 5, 10, 30, 60, 120):
                if base <= step:
                    col_sec = step
                    break
            else:
                col_sec = 120
        n_cols = int(math.ceil(max_end / col_sec)) if max_end > 0 else 0
        # 表头样式
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        # 时间块刻度表头（G 列起，F 列保留“层级”）
        for c in range(n_cols):
            t = (c + 1) * col_sec
            label = int(t) if col_sec >= 1 or abs(t - round(t)) < 1e-9 else round(t, 1)
            cell = ws.cell(row=1, column=7 + c, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        # 数据行块填充
        for i, r in enumerate(data, start=2):
            start_s = float(r[col_start] or 0.0)
            end_s = float(r[col_end] or r[col_start] or 0.0)
            layer = str(r[col_layer] or '循环')
            fill = layer_fill.get(layer, layer_fill['循环'])
            c0 = int(round(start_s / col_sec))
            c1 = max(int(round(end_s / col_sec)), c0 + 1)  # 至少 1 列，保证可见
            for c in range(c0, c1):
                if c >= n_cols:
                    break
                cell = ws.cell(row=i, column=7 + c)
                cell.fill = fill
                cell.border = border
            # 左侧数据单元格样式
            for j in range(min(6, len(r))):
                cell = ws.cell(row=i, column=j + 1, value=r[j])
                cell.alignment = center
                cell.border = border
        # 列宽：数据列自适应，时间块列固定小宽
        for j in range(1, 7):
            letter = get_column_letter(j)
            max_len = 0
            for rr in range(1, len(data) + 2):
                v = ws.cell(row=rr, column=j).value
                if v is None:
                    continue
                max_len = max(max_len, sum(2 if ord(ch) > 127 else 1 for ch in str(v)))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 8), 24)
        for c in range(1, n_cols + 1):
            ws.column_dimensions[get_column_letter(6 + c)].width = 2.5
        ws.freeze_panes = "F2"

    # ---------- 功能一：文档合并与内容拆分 ----------
    def process(self, source_dir, output_dir, keywords=None, separator=None,
                file_filters=None, rows=None, cancel_event=None, progress_callback=None,
                merge_groups=None, abnormal_keywords=None,
                step_units=None, step_mode=None, step_coefficient=1.5, step_max_seconds=None):
        if progress_callback:
            progress_callback(5)
        all_data = rows if rows is not None else self._read_all(
            source_dir, progress_callback, file_filters=file_filters, cancel_event=cancel_event)
        df_all = pd.DataFrame(all_data)

        if progress_callback:
            progress_callback(30)

        # 步骤超时异常行：先跑一次步骤分析，取异常触发行内容用于标红
        step_lines = set()
        if step_units or step_mode == 'sa':
            cutoff = step_max_seconds if step_max_seconds is not None else 900.0
            cached = getattr(self, '_step_lines_cache', None)
            if (cached and cached[0] == id(all_data)
                    and cached[1] == step_mode
                    and cached[2] == float(step_coefficient)
                    and cached[3] == cutoff):
                # 同一次运行中 UPH 步骤分析已算出异常行，直接复用避免重复扫描大日志
                step_lines = cached[4]
            else:
                if step_mode == 'sa':
                    sdf = analyze_steps_sa(all_data, step_coefficient,
                                           max_step_seconds=cutoff, cancel_event=cancel_event)
                elif step_units:
                    sdf = analyze_steps(all_data, step_units, step_coefficient,
                                        max_step_seconds=cutoff, cancel_event=cancel_event)
                step_lines = getattr(sdf, 'attrs', {}).get('anomaly_lines') or set()

        if merge_groups:
            # 按文件分组（如 PLC1/PLC2）分别导出 Excel
            return self._process_by_groups(
                all_data, output_dir, keywords, separator, merge_groups,
                cancel_event, progress_callback, abnormal_keywords, step_lines,
            )

        if len(all_data) > self.MERGE_MAX_TOTAL_ROWS:
            # 原始日志过大：放弃合并，按日志文件逐个导出 Excel
            return self._process_per_file(
                all_data, output_dir, keywords, separator, cancel_event, progress_callback,
                abnormal_keywords, step_lines,
            )

        df_all = pd.DataFrame(all_data)
        sheets = self._build_merge_sheets(df_all, keywords, separator, cancel_event)

        if progress_callback:
            progress_callback(70)

        out_path = os.path.join(output_dir, 'LogAnalysis.xlsx')
        hl = self._sheet_highlights(sheets, abnormal_keywords, step_lines, cancel_event)
        applied = self._write_sheets(out_path, sheets, cancel_event=cancel_event, sheet_highlights=hl)
        if not applied:
            self._apply_highlights(out_path, abnormal_keywords, step_lines, cancel_event)

        if progress_callback:
            progress_callback(100)
        return out_path

    def _process_by_groups(self, all_data, output_dir, keywords, separator, merge_groups,
                           cancel_event=None, progress_callback=None, abnormal_keywords=None,
                           step_lines=None):
        """按文件关键词分组分别导出 LogAnalysis_<组名>.xlsx，未匹配行归入“其他”。"""
        import os as _os
        _os.makedirs(output_dir, exist_ok=True)
        rows_list = list(all_data)
        written = []
        matched_ids = set()
        total = len(merge_groups)
        for gi, group in enumerate(merge_groups):
            if cancel_event is not None and cancel_event.is_set():
                raise OperationCancelled()
            name = group.get('name', '组%d' % (gi + 1))
            file_kw = group.get('file', '')
            sub = [r for r in rows_list if file_kw in str(r.get('FileName', ''))] if file_kw else rows_list
            if not sub:
                continue
            sheets = self._build_merge_sheets(pd.DataFrame(sub), keywords, separator, cancel_event)
            out_path = _os.path.join(output_dir, 'LogAnalysis_%s.xlsx' % name)
            hl = self._sheet_highlights(sheets, abnormal_keywords, step_lines, cancel_event)
            applied = self._write_sheets(out_path, sheets, cancel_event=cancel_event, sheet_highlights=hl)
            if not applied:
                self._apply_highlights(out_path, abnormal_keywords, step_lines, cancel_event)
            written.append(out_path)
            matched_ids.update(id(r) for r in sub)
            if progress_callback:
                progress_callback(20 + int(60 * (gi + 1) / total))
        others = [r for r in rows_list if id(r) not in matched_ids]
        if others:
            sheets = self._build_merge_sheets(pd.DataFrame(others), keywords, separator, cancel_event)
            out_path = _os.path.join(output_dir, 'LogAnalysis_其他.xlsx')
            hl = self._sheet_highlights(sheets, abnormal_keywords, step_lines, cancel_event)
            applied = self._write_sheets(out_path, sheets, cancel_event=cancel_event, sheet_highlights=hl)
            if not applied:
                self._apply_highlights(out_path, abnormal_keywords, step_lines, cancel_event)
            written.append(out_path)
        if progress_callback:
            progress_callback(100)
        if written:
            return "; ".join(written) + "（已按分组分别导出）"
        return "无匹配日志"

    @staticmethod
    def _highlight_abnormal_rows(path, abnormal_kws, cancel_event=None):
        """将内容命中异常关键词的日志行整行标红（浅红底 FFC7CE）。"""
        if not abnormal_kws:
            return
        kws = [k for k in re.split(r'[,，、;；\s]+', str(abnormal_kws)) if k]
        if not kws:
            return
        try:
            wb = load_workbook(path)
        except Exception:
            return
        red = PatternFill('solid', fgColor='FFC7CE')
        for ws in wb.worksheets:
            if ws.max_row < 2:
                continue
            header = {c.value: idx for idx, c in enumerate(ws[1], start=1)}
            content_col = header.get('Content') or header.get('OriginalContent')
            if not content_col:
                continue
            for row in ws.iter_rows(min_row=2):
                if cancel_event is not None and cancel_event.is_set():
                    break
                cell = row[content_col - 1]
                val = cell.value
                if val and any(kw in str(val) for kw in kws):
                    for c in row:
                        c.fill = red
        wb.save(path)

    def _apply_highlights(self, path, abnormal_keywords, step_lines, cancel_event=None):
        """合并输出统一标红：报警/停机关键词行 + UPH 步骤超时触发行。

        openpyxl 引擎兜底：单次加载、单次扫描、单次保存（xlsxwriter 引擎下
        写表阶段已用 set_row 完成标红，不再走到这里）。
        """
        if not abnormal_keywords and not step_lines:
            return
        pat = self._keyword_pattern(abnormal_keywords)
        if pat is None and not step_lines:
            return
        try:
            wb = load_workbook(path)
        except Exception:
            return
        red = PatternFill('solid', fgColor='FFC7CE')
        lines = set(step_lines or ())
        for ws in wb.worksheets:
            if ws.max_row < 2:
                continue
            header = {c.value: idx for idx, c in enumerate(ws[1], start=1)}
            content_col = header.get('Content') or header.get('OriginalContent')
            if not content_col:
                continue
            for row in ws.iter_rows(min_row=2):
                if cancel_event is not None and cancel_event.is_set():
                    break
                cell = row[content_col - 1]
                val = cell.value
                if val is None:
                    continue
                if (pat is not None and pat.search(str(val))) or (lines and val in lines):
                    for c in row:
                        c.fill = red
        wb.save(path)

    @staticmethod
    def _keyword_pattern(abnormal_keywords):
        """异常关键词 → 单条编译正则（IGNORECASE），供写表/兜底标红共用。"""
        if not abnormal_keywords:
            return None
        kws = [k for k in re.split(r'[,，、;；\s]+', str(abnormal_keywords)) if k]
        if not kws:
            return None
        return re.compile('|'.join(re.escape(k) for k in kws), re.IGNORECASE)

    @staticmethod
    def _compute_highlight_rows(df, abnormal_keywords, step_lines, cancel_event=None):
        """返回 DataFrame 中需标红的 0-based 行索引（异常关键词或步骤超时触发行）。

        用 pandas 向量化匹配（str.contains + isin），避免 openpyxl 逐格扫描大表。
        """
        if cancel_event is not None and cancel_event.is_set():
            raise OperationCancelled()
        col = 'Content' if 'Content' in df.columns else (
            'OriginalContent' if 'OriginalContent' in df.columns else None)
        if col is None or df.empty:
            return []
        mask = pd.Series(False, index=df.index)
        if abnormal_keywords:
            pat = LogModel._keyword_pattern(abnormal_keywords)
            if pat is not None:
                mask |= df[col].astype(str).str.contains(pat, na=False)
        if step_lines:
            mask |= df[col].isin(set(step_lines))
        return df.index[mask].tolist()

    def _sheet_highlights(self, sheets, abnormal_keywords, step_lines, cancel_event=None):
        """批量计算各 sheet 需标红的行索引；无关键词/无异常行时返回 None。"""
        if not abnormal_keywords and not step_lines:
            return None
        return {
            name: self._compute_highlight_rows(df, abnormal_keywords, step_lines, cancel_event)
            for name, df in sheets.items()
        }

    @staticmethod
    def _highlight_step_lines(path, lines, cancel_event=None):
        """将内容精确命中步骤超时异常触发行的日志行整行标红。"""
        if not lines:
            return
        try:
            wb = load_workbook(path)
        except Exception:
            return
        red = PatternFill('solid', fgColor='FFC7CE')
        for ws in wb.worksheets:
            if ws.max_row < 2:
                continue
            header = {c.value: idx for idx, c in enumerate(ws[1], start=1)}
            content_col = header.get('Content') or header.get('OriginalContent')
            if not content_col:
                continue
            for row in ws.iter_rows(min_row=2):
                if cancel_event is not None and cancel_event.is_set():
                    break
                cell = row[content_col - 1]
                if cell.value and cell.value in lines:
                    for c in row:
                        c.fill = red
        wb.save(path)

    def _build_merge_sheets(self, df_all, keywords, separator, cancel_event=None):
        """关键词筛选 + 分隔符拆分，返回 {'AllLogs':..., 'Filtered':...}。"""
        if keywords:
            kw_list = re.split(r'[ ;、，]+', keywords)
            kw_list = [k for k in kw_list if k]
            if kw_list:
                pattern = '|'.join(re.escape(k) for k in kw_list)
                mask = df_all['Content'].str.contains(pattern, case=False, na=False)
                filtered_df = df_all[mask].copy()
            else:
                filtered_df = df_all.copy()
        else:
            filtered_df = df_all.copy()

        has_filtered_rows = not filtered_df.empty
        has_separator = separator is not None and separator != ''
        has_keywords = bool(keywords and str(keywords).strip())
        # 无关键词且无分隔符时 Filtered 与 AllLogs 完全相同，跳过以加速大文件导出
        generate_filtered = (has_filtered_rows or has_separator) and (has_keywords or has_separator)

        if generate_filtered:
            if not has_filtered_rows and has_separator:
                filtered_df = df_all.copy()
                has_filtered_rows = True
            if has_separator:
                sep = separator
                if sep == '\\t':
                    sep = '\t'
                split_rows = []
                max_parts = 0
                for _, row in filtered_df.iterrows():
                    if cancel_event is not None and cancel_event.is_set():
                        raise OperationCancelled()
                    parts = row['Content'].split(sep)
                    parts = [clean_for_excel(p) for p in parts]
                    max_parts = max(max_parts, len(parts))
                    split_rows.append(parts)
                col_names = [f'Part_{i+1}' for i in range(max_parts)]
                split_data = []
                for parts in split_rows:
                    parts += [''] * (max_parts - len(parts))
                    split_data.append(parts)
                split_df = pd.DataFrame(split_data, columns=col_names)
                filtered_df = filtered_df.reset_index(drop=True)
                filtered_df = pd.concat([
                    filtered_df[['FileName']],
                    pd.DataFrame({'OriginalContent': filtered_df['Content']}),
                    split_df
                ], axis=1)
                # Content 已由 OriginalContent 取代，无需（也无法）再删除

        sheets = {'AllLogs': df_all}
        if generate_filtered and filtered_df is not None and not filtered_df.empty:
            sheets['Filtered'] = filtered_df
        return sheets

    def _process_per_file(self, all_data, output_dir, keywords, separator, cancel_event=None,
                          progress_callback=None, abnormal_keywords=None, step_lines=None):
        """日志过大：新建子文件夹，按日志文件逐个导出 Excel（文件名 = 日志文件名去掉扩展名）。"""
        sub_dir = os.path.join(output_dir, 'LogAnalysis_Files')
        os.makedirs(sub_dir, exist_ok=True)
        has_keywords = bool(keywords and str(keywords).strip())
        has_separator = separator is not None and separator != ''
        if not has_keywords and not has_separator:
            # 无筛选：直接流式写入，避免构建大 DataFrame（最快、内存最低）
            n = self._stream_per_file(
                sub_dir, all_data, cancel_event,
                abnormal_keywords=abnormal_keywords, step_lines=step_lines,
            )
            if progress_callback:
                progress_callback(100)
            return f"{sub_dir}（日志过大，已按文件分别导出 {n} 个 Excel）"

        file_names = sorted({r.get('FileName', '') for r in all_data})
        n = len(file_names)
        used_names = set()
        for idx, fname in enumerate(file_names):
            if cancel_event is not None and cancel_event.is_set():
                raise OperationCancelled()
            sub = pd.DataFrame([r for r in all_data if r.get('FileName') == fname])
            sheets = self._build_merge_sheets(sub, keywords, separator, cancel_event)
            out_path = os.path.join(sub_dir, self._per_file_output_name(fname, used_names))
            hl = self._sheet_highlights(sheets, abnormal_keywords, step_lines, cancel_event)
            applied = self._write_sheets(out_path, sheets, cancel_event=cancel_event, sheet_highlights=hl)
            if not applied:
                self._apply_highlights(out_path, abnormal_keywords, step_lines, cancel_event)
            if progress_callback:
                progress_callback(30 + int(60 * (idx + 1) / n))
        if progress_callback:
            progress_callback(100)
        return f"{sub_dir}（日志过大，已按文件分别导出 {n} 个 Excel）"

    @staticmethod
    def _stream_per_file(sub_dir, all_data, cancel_event=None,
                         abnormal_keywords=None, step_lines=None):
        """无筛选时按日志文件流式写入 Excel（xlsxwriter 常量内存模式），写表同时按行标红。返回文件数。"""
        import xlsxwriter
        max_rows = 1048575
        # 预计算需标红的内容集合（pandas 向量化，避免写表时逐行 Python 正则）
        hl_set = set()
        pat = LogModel._keyword_pattern(abnormal_keywords)
        lines = set(step_lines or ())
        if pat is not None or lines:
            s = pd.Series([str(r.get('Content', '')) for r in all_data])
            mask = pd.Series(False, index=s.index)
            if pat is not None:
                mask |= s.str.contains(pat, na=False)
            if lines:
                mask |= s.isin(lines)
            hl_set = set(s[mask].tolist())
        # 按文件分组（一次遍历）。随后每个文件独立建表、写完即关，
        # 避免同时打开几百个 workbook 的临时文件句柄（Windows 默认约 512 个上限）。
        by_file = {}
        for row in all_data:
            if cancel_event is not None and cancel_event.is_set():
                raise OperationCancelled()
            by_file.setdefault(row.get('FileName', ''), []).append(row)
        n = 0
        used_names = set()
        for fname in sorted(by_file):
            if cancel_event is not None and cancel_event.is_set():
                raise OperationCancelled()
            out_name = LogModel._per_file_output_name(fname, used_names)
            wb = xlsxwriter.Workbook(
                os.path.join(sub_dir, out_name),
                {'nan_inf_to_errors': True, 'constant_memory': True},
            )
            ws = wb.add_worksheet('AllLogs')
            ws.write_string(0, 0, 'FileName')
            ws.write_string(0, 1, 'Content')
            red = wb.add_format({'bg_color': 'FFC7CE'}) if hl_set else None
            r = 1
            sheet_no = 1
            for row in by_file[fname]:
                if cancel_event is not None and cancel_event.is_set():
                    raise OperationCancelled()
                if r > max_rows:
                    sheet_no += 1
                    ws = wb.add_worksheet(f'AllLogs_{sheet_no}'[:31])
                    ws.write_string(0, 0, 'FileName')
                    ws.write_string(0, 1, 'Content')
                    r = 1
                content = str(row.get('Content', ''))
                if hl_set and content in hl_set:
                    # xlsxwriter 常量内存模式：行格式必须在写该行之前设置
                    ws.set_row(r, None, red)
                ws.write_string(r, 0, str(fname))
                ws.write_string(r, 1, content)
                r += 1
            wb.close()
            n += 1
        return n

    @staticmethod
    def _per_file_output_name(fname, used_names):
        """按日志文件名生成导出 Excel 文件名。

        同名不同扩展名的日志（如 SA 的 xxx.log 与 xxx.txt 是同一数据两种格式）
        会映射到同一 xlsx 路径导致互相覆盖；撞车时追加原扩展名区分。
        """
        stem = os.path.splitext(fname)[0].replace('/', '_').replace('\\', '_')
        name = f'{stem}.xlsx'
        if name in used_names:
            ext = os.path.splitext(fname)[1].lstrip('.') or 'log'
            name = f'{stem}_{ext}.xlsx'
        used_names.add(name)
        return name

    # ---------- 功能二：UPH 分析 ----------
    def analyze_uph(self, source_dir, output_dir, trigger_keywords="MarkEnd1", units_per_cycle=1,
                    normal_threshold=10.0, planned_threshold=900.0,
                    ideal_ct=None, max_ct=None, file_filters=None, module_pattern=None,
                    pure_uph_factor=1.0, bottleneck_stations=None, bottleneck_units_per_row=None,
                    bottleneck_machines=None, tray_change=None, parts=None, module_from_path=False,
                    step_units=None, step_coefficient=1.5, step_max_seconds=None, step_mode=None,
                    rows=None, cancel_event=None, progress_callback=None):
        if progress_callback:
            progress_callback(5)
        rows = rows if rows is not None else self._read_all(
            source_dir, progress_callback, file_filters=file_filters, cancel_event=cancel_event)
        if progress_callback:
            progress_callback(40)

        steps_df = None
        gantt_df = None
        if step_units or step_mode == 'sa':
            cutoff = step_max_seconds if step_max_seconds is not None else planned_threshold
            if step_mode == 'sa':
                steps_df = analyze_steps_sa(
                    rows, step_coefficient,
                    max_step_seconds=cutoff, cancel_event=cancel_event,
                )
            elif step_units:
                steps_df = analyze_steps(
                    rows, step_units, step_coefficient,
                    max_step_seconds=cutoff, cancel_event=cancel_event,
                )
            # 缓存步骤超时异常行：同一运行内后续“文档合并与内容拆分”标红时直接复用，
            # 避免对 400 万行级日志重复跑一遍步骤分析
            self._step_lines_cache = (
                id(rows), step_mode, float(step_coefficient), cutoff,
                getattr(steps_df, 'attrs', {}).get('anomaly_lines') or set(),
            )
        if step_mode == 'sa':
            gantt_df = build_gantt_rows_sa(
                rows, max_step_seconds=cutoff, cancel_event=cancel_event)
        elif step_units:
            gantt_df = build_gantt_rows(
                rows, step_units, max_step_seconds=cutoff, cancel_event=cancel_event)
        if bottleneck_machines:
            return self._analyze_uph_machines(
                rows, bottleneck_machines, steps_df, gantt_df,
                output_dir, cancel_event, progress_callback)
        if parts:
            return self._analyze_uph_parts(
                rows, parts, units_per_cycle, normal_threshold, planned_threshold,
                ideal_ct, max_ct, pure_uph_factor, module_from_path,
                steps_df, gantt_df, output_dir, cancel_event, progress_callback)
        if bottleneck_stations:
            return self._analyze_uph_stations(
                rows, bottleneck_stations, bottleneck_units_per_row or units_per_cycle,
                tray_change, steps_df, gantt_df,
                output_dir, cancel_event, progress_callback)
        return self._analyze_uph_basic(
            rows, trigger_keywords, units_per_cycle, normal_threshold, planned_threshold,
            ideal_ct, max_ct, module_pattern, pure_uph_factor, tray_change,
            steps_df, gantt_df, output_dir, cancel_event, progress_callback)

    def _analyze_uph_machines(self, rows, bottleneck_machines, steps_df, gantt_df, output_dir, cancel_event, progress_callback):

        # CAW 双机台：上料机/焊接机 各自单颗 CT，取 CT 长者计算 UPH
        machine_df, bn, cycles_df = analyze_bottleneck_machines(
            rows, bottleneck_machines, cancel_event=cancel_event,
        )
        em = parse_em_production(rows, cancel_event=cancel_event)
        ame = pd.DataFrame([{
            '瓶颈机台': bn.get('瓶颈机台', ''),
            '瓶颈CT(秒)': bn.get('瓶颈CT(秒)', ''),
            'UPH(个/小时)': bn.get('UPH(个/小时)', ''),
        }])
        sheets = {'Summary': machine_df, 'AMESummary': ame}
        if cycles_df is not None and not cycles_df.empty:
            sheets['CycleDetail'] = cycles_df
        if steps_df is not None and not steps_df.empty:
            sheets['步骤分析'] = steps_df
        if gantt_df is not None and not gantt_df.empty:
            sheets['步骤甘特图'] = gantt_df
        if not em.empty:
            sheets['EMProduction'] = em
        if progress_callback:
            progress_callback(70)
        out_path = os.path.join(output_dir, 'UPH_Analysis.xlsx')
        self._write_sheets(out_path, sheets, cancel_event=cancel_event)
        if progress_callback:
            progress_callback(100)
        return out_path


    def _analyze_uph_parts(self, rows, parts, units_per_cycle, normal_threshold, planned_threshold, ideal_ct, max_ct, pure_uph_factor, module_from_path, steps_df, gantt_df, output_dir, cancel_event, progress_callback):

        # 多部分机台（如上料机/主机/下料机）：各部分独立 UPH，换盘时间可平摊
        part_summaries = []
        ame_rows = []
        for part in parts:
            name = part.get('name', '')
            trigger = part.get('trigger', '')
            units = int(part.get('units_per_cycle', units_per_cycle or 1))
            part_normal = float(part.get('normal_threshold', normal_threshold))
            part_planned = float(part.get('planned_threshold', planned_threshold))
            p_rows = [r for r in rows if str(r.get('FileName', '')).split('/')[0] == name]
            cycles = build_cycles_df(p_rows, trigger, part_normal, part_planned,
                                     module_from_path=module_from_path, cancel_event=cancel_event)
            if cycles.empty:
                row = {'模块': name, '周期总数': 0, '产出数(个)': 0, '统计时长(秒)': 0,
                       'UPH(个/小时)': '', 'Pure UPH(个/小时)': '', 'Derated UPH M2(个/小时)': ''}
                if part.get('tray_seconds') or part.get('tray_detect'):
                    row.update({'每盘颗数(统计)': '', '换盘次数': '',
                                '单次换盘时间(秒)': '', '每颗换盘开销(秒)': '',
                                '有效周期(秒)': '', '有效UPH(个/小时)': ''})
                part_summaries.append(pd.DataFrame([row]))
                ame_rows.append({'模块': name, 'Pure UPH(个/小时)': '', 'UPH(个/小时)': '', '产出数(个)': 0})
                continue
            s = summarize_uph(cycles, units, ideal_ct=ideal_ct, max_ct=max_ct,
                              pure_uph_factor=pure_uph_factor)
            tray_s = part.get('tray_seconds')
            units_per_tray = part.get('units_per_tray') or part.get('rows_per_tray')
            tray_stats = None
            if part.get('tray_detect'):
                try:
                    tray_stats = detect_tray_stats(
                        p_rows,
                        part['tray_detect']['tray_id'],
                        part['tray_detect'].get('unit'),
                        segments=part['tray_detect'].get('segments', 'id'),
                        run_gap=float(part['tray_detect'].get('run_gap', 300.0)),
                        cancel_event=cancel_event,
                    )
                except Exception:
                    tray_stats = None
            if tray_stats:
                units_per_tray = tray_stats['units_per_tray']
                s['每盘颗数(统计)'] = tray_stats['units_per_tray']
                s['换盘次数'] = tray_stats['tray_count']
                s['单次换盘时间(秒)'] = tray_stats['tray_seconds']
            if part.get('tray_change'):
                # SA 式：换盘时间 = 同工位 卸载 -> 下一次装载 的间隔
                tc = measure_tray_change(
                    p_rows,
                    part['tray_change'].get('unload', ''),
                    part['tray_change'].get('load', ''),
                    cancel_event=cancel_event,
                )
                if tc:
                    tray_s = tc['tray_seconds']
                    s['换盘次数'] = tc['tray_count']
                    s['单次换盘时间(秒)'] = tc['tray_seconds']
            if tray_s and units_per_tray:
                try:
                    pure_f = float(s.iloc[0].get('Pure UPH(个/小时)'))
                    base_cycle = 3600.0 * units / pure_f
                    overhead = float(tray_s) / float(units_per_tray)
                    eff_cycle = base_cycle + overhead
                    s['每颗换盘开销(秒)'] = round(overhead, 3)
                    s['有效周期(秒)'] = round(eff_cycle, 3)
                    s['有效UPH(个/小时)'] = round(3600.0 * units / eff_cycle, 2)
                except (TypeError, ValueError, ZeroDivisionError):
                    pass
            part_summaries.append(s)
            r0 = s.iloc[0]
            p_em = parse_em_production(p_rows, cancel_event=cancel_event)
            em_input = int(p_em['InputQty'].sum()) if not p_em.empty else ''
            em_good = int(p_em['GoodQty'].sum()) if not p_em.empty else ''
            ame_rows.append({
                '模块': name,
                'Pure UPH(个/小时)': r0.get('Pure UPH(个/小时)'),
                'Derated UPH M1(个/小时)': '',
                'Derated UPH M2(个/小时)': r0.get('Derated UPH M2(个/小时)'),
                'UPH(个/小时)': r0.get('UPH(个/小时)'),
                '产出数(个)': r0.get('产出数(个)'),
                'EM投入数(个)': em_input,
                'EM良品数(个)': em_good,
                '运行时间RUN(秒)': '',
                '周期总数': r0.get('周期总数'),
                '统计时长(秒)': r0.get('统计时长(秒)'),
                '平均正常周期(秒)': r0.get('平均正常周期(秒)'),
            })
        summary = pd.concat(part_summaries, ignore_index=True)
        ame_summary = pd.DataFrame(ame_rows)
        em_all = parse_em_production(rows, cancel_event=cancel_event)
        if progress_callback:
            progress_callback(70)
        sheets = {'Summary': summary, 'AMESummary': ame_summary}
        if not em_all.empty:
            sheets['EMProduction'] = em_all
        if steps_df is not None and not steps_df.empty:
            sheets['步骤分析'] = steps_df
        if gantt_df is not None and not gantt_df.empty:
            sheets['步骤甘特图'] = gantt_df
        out_path = os.path.join(output_dir, 'UPH_Analysis.xlsx')
        self._write_sheets(out_path, sheets, cancel_event=cancel_event)
        if progress_callback:
            progress_callback(100)
        return out_path


    def _analyze_uph_stations(self, rows, bottleneck_stations, units, tray_change, steps_df, gantt_df, output_dir, cancel_event, progress_callback):

        # 多工位机：自动判定瓶颈工位，UPH = 每排产品数 × 3600 / 瓶颈工位每排周期
        stations_df, bn = analyze_bottleneck(
            rows, bottleneck_stations, units, tray_change=tray_change,
            cancel_event=cancel_event,
        )
        b_name = bn.get('瓶颈工位') or ''
        eff_cycle = bn.get('有效周期(秒)')
        pure = round(units * 3600.0 / eff_cycle, 2) if eff_cycle else ''
        em = parse_em_production(rows, cancel_event=cancel_event)
        ame = pd.DataFrame([{
            '瓶颈工位': b_name,
            '瓶颈周期(秒)': bn.get('瓶颈周期(秒)') or '',
            '换盘次数': bn.get('换盘次数') or '',
            '单次换盘时间(秒)': bn.get('单次换盘时间(秒)') or '',
            '每盘排数': bn.get('每盘排数') or '',
            '每排换盘开销(秒)': bn.get('每排换盘开销(秒)') or '',
            '有效周期(秒)': eff_cycle if eff_cycle else '',
            '单颗CT(秒)': round(eff_cycle / units, 3) if eff_cycle else '',
            '每排产品数(个)': units,
            'Pure UPH(个/小时)': pure,
            'Derated UPH M1(个/小时)': '',
            'Derated UPH M2(个/小时)': pure if pure != '' else '',
            'EM投入数(个)': int(em['InputQty'].sum()) if not em.empty else '',
        }])
        if progress_callback:
            progress_callback(70)
        sheets = {'Summary': stations_df, 'AMESummary': ame}
        if not em.empty:
            sheets['EMProduction'] = em
        if steps_df is not None and not steps_df.empty:
            sheets['步骤分析'] = steps_df
        if gantt_df is not None and not gantt_df.empty:
            sheets['步骤甘特图'] = gantt_df
        out_path = os.path.join(output_dir, 'UPH_Analysis.xlsx')
        self._write_sheets(out_path, sheets, cancel_event=cancel_event)
        if progress_callback:
            progress_callback(100)
        return out_path

    def _analyze_uph_basic(self, rows, trigger_keywords, units_per_cycle, normal_threshold, planned_threshold, ideal_ct, max_ct, module_pattern, pure_uph_factor, tray_change, steps_df, gantt_df, output_dir, cancel_event, progress_callback):
        cycles = build_cycles_df(rows, trigger_keywords, normal_threshold, planned_threshold,
                                 module_pattern=module_pattern, cancel_event=cancel_event)
        # 换盘时间平摊到整盘产品（LM：整盘打标结束 下料->新盘上料，按 CCD 批次颗数×每盘批数 得每盘颗数）
        tray_overhead = None
        tray_cols = {}
        if tray_change:
            tc = measure_tray_change(
                rows, tray_change.get('unload', ''), tray_change.get('load', ''),
                cancel_event=cancel_event,
            )
            det = None
            if tray_change.get('batch') and tray_change.get('unit'):
                det = detect_units_per_tray(
                    rows, tray_change['batch'], tray_change['unit'],
                    tray_change.get('tray', ''), cancel_event=cancel_event,
                )
            units_per_tray = det.get('units_per_tray') if det else None
            if tc and units_per_tray:
                tray_overhead = tc['tray_seconds'] / float(units_per_tray)
                tray_cols = {
                    '每盘颗数(统计)': units_per_tray,
                    '每批颗数(统计)': det.get('units_per_batch'),
                    '换盘次数': tc['tray_count'],
                    '单次换盘时间(秒)': tc['tray_seconds'],
                    '每颗换盘开销(秒)': round(tray_overhead, 4),
                }
        summary = summarize_uph(cycles, units_per_cycle, ideal_ct=ideal_ct, max_ct=max_ct,
                                pure_uph_factor=pure_uph_factor,
                                tray_overhead_seconds=tray_overhead)
        if tray_cols:
            for k, v in tray_cols.items():
                summary[k] = v
        em = parse_em_production(rows, cancel_event=cancel_event)
        status_summary, _, _ = analyze_status(rows, cancel_event=cancel_event)
        run_seconds = None
        if not status_summary.empty:
            run_rows = status_summary.loc[status_summary['状态'] == 'RUN', '总时长(秒)']
            run_seconds = float(run_rows.sum()) if len(run_rows) else None
        ame_summary = summarize_uph_ame(
            cycles, units_per_cycle, ideal_ct=ideal_ct, max_ct=max_ct,
            em_df=em, run_seconds=run_seconds,
            tray_overhead_seconds=tray_overhead,
        )
        if tray_cols:
            for k, v in tray_cols.items():
                ame_summary[k] = v
        if progress_callback:
            progress_callback(70)
        sheets = {'Summary': summary, 'AMESummary': ame_summary}
        if not cycles.empty:
            sheets['CycleDetail'] = cycles
        if not em.empty:
            sheets['EMProduction'] = em
        if steps_df is not None and not steps_df.empty:
            sheets['步骤分析'] = steps_df
        if gantt_df is not None and not gantt_df.empty:
            sheets['步骤甘特图'] = gantt_df
        out_path = os.path.join(output_dir, 'UPH_Analysis.xlsx')
        self._write_sheets(out_path, sheets, cancel_event=cancel_event)
        if progress_callback:
            progress_callback(100)
        return out_path



    # ---------- 功能三：EFF 分析 ----------
    def analyze_eff(self, source_dir, output_dir, planned_hours=None,
                    pdt_reason_ids=None, reason_device=None, file_filters=None, rows=None,
                    cancel_event=None, activity_keywords=None, stop_reason_keywords=None,
                    progress_callback=None):
        """CoreTech AME 效率：EFF = 操作时间(运行+待机) / 计划生产时间，基于 RUN/IDLE/DOWN 状态。"""
        reason_map = LogModel._load_reason_map(reason_device)
        if progress_callback:
            progress_callback(5)
        rows = rows if rows is not None else self._read_all(
            source_dir, progress_callback, file_filters=file_filters, cancel_event=cancel_event)
        if progress_callback:
            progress_callback(50)
        if activity_keywords:
            status_summary, hourly, detail = analyze_status_derived(
                rows, activity_keywords, stop_reason_keywords,
                reason_map=reason_map, cancel_event=cancel_event,
            )
        else:
            status_summary, hourly, detail = analyze_status(rows, cancel_event=cancel_event)
        summary = summarize_eff_coretech(
            status_summary, detail, planned_hours=planned_hours, pdt_reason_ids=pdt_reason_ids,
            reason_map=reason_map,
        )
        pareto = down_pareto(detail, reason_map=reason_map)
        if progress_callback:
            progress_callback(70)
        sheets = {'Summary': summary}
        if not hourly.empty:
            sheets['Hourly'] = hourly
        if not pareto.empty:
            sheets['DOWN_Pareto'] = pareto
        if not detail.empty:
            sheets['Detail'] = detail
        out_path = os.path.join(output_dir, 'EFF_Analysis.xlsx')
        self._write_sheets(out_path, sheets, cancel_event=cancel_event)
        if progress_callback:
            progress_callback(100)
        return out_path

    # ---------- 功能四：报警分析 ----------
    def analyze_alarms(self, source_dir, output_dir, alarm_keywords="报警,ALARM,ERROR,NG,失败,异常,停止信号",
                       file_filters=None, rows=None, cancel_event=None, reason_device=None,
                       module_from_path=False, progress_callback=None):
        reason_map = LogModel._load_reason_map(reason_device)
        if progress_callback:
            progress_callback(5)
        rows = rows if rows is not None else self._read_all(
            source_dir, progress_callback, file_filters=file_filters, cancel_event=cancel_event)
        if progress_callback:
            progress_callback(50)
        summary, by_keyword, detail = summarize_alarms(
            rows, alarm_keywords, cancel_event=cancel_event, reason_map=reason_map,
            module_from_path=module_from_path,
        )
        if progress_callback:
            progress_callback(70)
        sheets = {'Summary': summary}
        if not by_keyword.empty:
            sheets['ByKeyword'] = by_keyword
        if not detail.empty:
            sheets['Detail'] = detail
        out_path = os.path.join(output_dir, 'Alarm_Analysis.xlsx')
        self._write_sheets(out_path, sheets, cancel_event=cancel_event)
        if progress_callback:
            progress_callback(100)
        return out_path

    # ---------- 功能五：机台状态分析 ----------
    def analyze_status(self, source_dir, output_dir, file_filters=None, rows=None, cancel_event=None,
                       activity_keywords=None, stop_reason_keywords=None, reason_device=None,
                       progress_callback=None):
        reason_map = LogModel._load_reason_map(reason_device)
        if progress_callback:
            progress_callback(5)
        rows = rows if rows is not None else self._read_all(
            source_dir, progress_callback, file_filters=file_filters, cancel_event=cancel_event)
        if progress_callback:
            progress_callback(50)
        if activity_keywords:
            summary, hourly, detail = analyze_status_derived(
                rows, activity_keywords, stop_reason_keywords,
                reason_map=reason_map, cancel_event=cancel_event,
            )
        else:
            summary, hourly, detail = analyze_status(rows, cancel_event=cancel_event)
        if progress_callback:
            progress_callback(70)
        sheets = {'Summary': summary}
        if not hourly.empty:
            sheets['Hourly'] = hourly
        if not detail.empty:
            sheets['Detail'] = detail
        out_path = os.path.join(output_dir, 'Status_Analysis.xlsx')
        self._write_sheets(out_path, sheets, cancel_event=cancel_event)
        if progress_callback:
            progress_callback(100)
        return out_path
