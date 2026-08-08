import os
import tempfile
import unittest

import pandas as pd

from models.log_model import LogModel


class LogModelTest(unittest.TestCase):
    def setUp(self):
        self.src = tempfile.mkdtemp()
        self.out = tempfile.mkdtemp()
        with open(os.path.join(self.src, "a.log"), "w", encoding="utf-8") as f:
            f.write("2026-08-05 10:00:01 ERROR: servo timeout\n")
            f.write("2026-08-05 10:00:02 INFO: motion done\n")
        with open(os.path.join(self.src, "b.txt"), "w", encoding="gbk") as f:
            f.write("2026-08-05 10:00:03 ERROR: axis jammed\n")

    def _parse(self, **kwargs):
        return LogModel().process(self.src, self.out, **kwargs)

    def test_no_separator_filter(self):
        path = self._parse(keywords="ERROR")
        xl = pd.ExcelFile(path)
        self.assertEqual(xl.sheet_names, ["AllLogs", "Filtered"])
        self.assertEqual(len(xl.parse("AllLogs")), 3)
        self.assertEqual(len(xl.parse("Filtered")), 2)

    def test_separator_with_keywords(self):
        path = self._parse(keywords="ERROR", separator=",")
        df = pd.ExcelFile(path).parse("Filtered")
        self.assertIn("OriginalContent", df.columns)
        self.assertNotIn("Content", df.columns)

    def test_separator_without_keywords(self):
        path = self._parse(separator=",")
        df = pd.ExcelFile(path).parse("Filtered")
        self.assertIn("Part_1", df.columns)
        self.assertNotIn("Content", df.columns)

    def test_progress_callback_reaches_100(self):
        seen = []
        self._parse(keywords="ERROR", progress_callback=seen.append)
        self.assertEqual(seen[-1], 100)

    def test_read_files_recursive(self):
        from utils.file_utils import read_files
        src = tempfile.mkdtemp()
        sub = os.path.join(src, "day1", "station1")
        os.makedirs(sub)
        with open(os.path.join(sub, "a.log"), "w", encoding="utf-8") as f:
            f.write("2026-07-08 00:00:01.000 [status:RUN]\n")
        with open(os.path.join(src, "b.txt"), "w", encoding="utf-8") as f:
            f.write("hello\n")
        rows = read_files(src)
        names = {r["FileName"] for r in rows}
        self.assertIn("day1/station1/a.log", names)
        self.assertIn("b.txt", names)

    def test_read_files_file_filters(self):
        from utils.file_utils import read_files
        src = tempfile.mkdtemp()
        with open(os.path.join(src, "RAYPRUS交互记录.log"), "w", encoding="utf-8") as f:
            f.write("2026-07-08 00:00:01.000 x\n")
        with open(os.path.join(src, "记录PLC2当前工位当前步数记录.log"), "w", encoding="utf-8") as f:
            f.write("2026-07-08 00:00:02.000 y\n")
        with open(os.path.join(src, "Debug记录.log"), "w", encoding="utf-8") as f:
            f.write("2026-07-08 00:00:03.000 z\n")
        rows = read_files(src, file_filters=["记录PLC2", "RAYPRUS"])
        names = {r["FileName"] for r in rows}
        self.assertEqual(names, {"记录PLC2当前工位当前步数记录.log", "RAYPRUS交互记录.log"})

    def test_write_sheets_split_large(self):
        big = pd.DataFrame({"a": range(1048575 + 10)})
        out = os.path.join(tempfile.mkdtemp(), "big.xlsx")
        LogModel._write_sheets(out, {"AllLogs": big})
        xl = pd.ExcelFile(out)
        self.assertEqual(xl.sheet_names, ["AllLogs_1", "AllLogs_2"])
        self.assertEqual(len(xl.parse("AllLogs_1")), 1048575)
        self.assertEqual(len(xl.parse("AllLogs_2")), 10)
        from openpyxl import load_workbook
        ws = load_workbook(out)["AllLogs_1"]
        self.assertNotIn("DDEBF7", str(ws.cell(row=1, column=1).fill.start_color.rgb))  # 超大表跳过格式化

    def test_process_per_file_when_too_large(self):
        rows = []
        for fname, cnt in (("a.log", 4), ("b.log", 4)):
            for i in range(cnt):
                rows.append({"FileName": fname, "Content": f"2026-07-08 00:00:0{i} x"})
        out = tempfile.mkdtemp()
        old = LogModel.MERGE_MAX_TOTAL_ROWS
        LogModel.MERGE_MAX_TOTAL_ROWS = 5  # 共 8 行 > 5 → 按文件导出
        try:
            result = LogModel().process("不存在的目录", out, rows=rows)
        finally:
            LogModel.MERGE_MAX_TOTAL_ROWS = old
        self.assertIn("已按文件分别导出 2 个 Excel", result)
        sub = os.path.join(out, "LogAnalysis_Files")
        self.assertTrue(os.path.isdir(sub))
        self.assertTrue(os.path.exists(os.path.join(sub, "a.xlsx")))
        self.assertTrue(os.path.exists(os.path.join(sub, "b.xlsx")))
        self.assertFalse(os.path.exists(os.path.join(out, "LogAnalysis.xlsx")))
        xl = pd.ExcelFile(os.path.join(sub, "a.xlsx"))
        self.assertEqual(xl.sheet_names, ["AllLogs"])
        self.assertEqual(len(xl.parse("AllLogs")), 4)

    def test_process_merged_when_small(self):
        rows = [{"FileName": "a.log", "Content": "2026-07-08 00:00:00 x"}]
        out = tempfile.mkdtemp()
        path = LogModel().process("不存在的目录", out, rows=rows)
        self.assertTrue(os.path.exists(path))  # LogAnalysis.xlsx

    def test_detect_encoding_gbk_single_byte_fallback(self):
        # 验证 GBK 中文日志能正确解码（不依赖 chardet 对编码标签的具体命名，
        # 不同 chardet 版本可能返回 GB2312/GBK/GB18030，标签本身不影响解码正确性）
        from utils.file_utils import read_file_with_fallback
        src = tempfile.mkdtemp()
        path = os.path.join(src, "gbk.log")
        with open(path, "wb") as f:
            f.write("2026-07-08 00:00:00 真空信号不达标\n".encode("gbk"))
        lines, enc = read_file_with_fallback(path)
        self.assertEqual(lines, ["2026-07-08 00:00:00 真空信号不达标"])
        self.assertIn(enc.lower().replace("-", "").replace("_", ""),
                      ("gbk", "gb2312", "gb18030"))

    def test_find_external_resource(self):
        from utils.resource_utils import find_external_resource
        path = find_external_resource("Analysis_Report.pptx")
        if not path:
            self.skipTest("Analysis_Report.pptx 为内部数据未入库（CI 环境缺失），跳过")
        self.assertTrue(os.path.exists(path))

    def test_find_external_resource_frozen(self):
        import sys as _sys
        from utils import resource_utils
        tmp = tempfile.mkdtemp()
        fake_exe = os.path.join(tmp, "MC_LogAnalysis.exe")
        with open(fake_exe, "wb") as f:
            f.write(b"")
        with open(os.path.join(tmp, "Analysis_Report.pptx"), "wb") as f:
            f.write(b"x")
        old_frozen = getattr(_sys, "frozen", None)
        old_exe = _sys.executable
        _sys.frozen = True
        _sys.executable = fake_exe
        try:
            found = resource_utils.find_external_resource("Analysis_Report.pptx")
            self.assertEqual(found, os.path.join(tmp, "Analysis_Report.pptx"))
        finally:
            if old_frozen is None:
                del _sys.frozen
            else:
                _sys.frozen = old_frozen
            _sys.executable = old_exe

    def test_excel_formatting(self):
        out = os.path.join(tempfile.mkdtemp(), "fmt.xlsx")
        LogModel._write_sheets(out, {"S": pd.DataFrame({"名称": ["A", "B"], "数量": [1, ""]})})
        from openpyxl import load_workbook
        ws = load_workbook(out)["S"]
        header = ws.cell(row=1, column=1)
        self.assertIn("DDEBF7", str(header.fill.start_color.rgb))  # 表头浅蓝
        self.assertTrue(header.font.bold)                           # 表头加粗
        content = ws.cell(row=2, column=1)
        self.assertEqual(content.border.left.style, "thin")        # 有内容加框线
        self.assertEqual(content.alignment.horizontal, "center")   # 内容居中
        empty = ws.cell(row=3, column=2)
        self.assertIsNone(empty.border.left.style)                 # 空单元格无格线
        self.assertGreaterEqual(ws.column_dimensions['A'].width, 6)  # 列宽自适应


if __name__ == "__main__":
    unittest.main()
